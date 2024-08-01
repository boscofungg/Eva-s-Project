import re
import openpyxl

def remove_brackets(input_string):
    """
    Remove all bracketed content (square, round, and Chinese-style) from the input string.
    """
    # Remove content in square brackets
    output_string = re.sub(r'\[[^\]]*\]', '', input_string)
    
    # Remove content in round brackets
    output_string = re.sub(r'\([^\)]*\)', '', output_string)
    
    # Remove content in Chinese-style brackets
    output_string = re.sub(r'（[^）]*）', '', output_string)
    
    return output_string

def remove_brackets_in_excel(input_file, output_file):
    """
    Remove all bracketed content from an Excel file.
    """
    # Load the workbook
    workbook = openpyxl.load_workbook(input_file)
    worksheet = workbook.active
    
    # Iterate through each cell and remove bracketed content
    for row in range(1, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row, column=col)
            if cell.value:
                new_value = remove_brackets(str(cell.value))
                new_value = new_value.strip()
                cell.value = new_value
    
    # Save the modified workbook
    workbook.save(output_file)

# Example usage
remove_brackets_in_excel('contract_1.xlsx', 'output.xlsx')
