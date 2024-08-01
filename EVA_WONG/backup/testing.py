import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from tkinter import simpledialog
import re
import pandas as pd
import openpyxl
import re


def get_user_input():
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    
    # Get user input using a simple dialog
    user_input = simpledialog.askstring("User Input", "Which contract are you working on now?")
    
    if user_input is not None:
        print("Contract:", user_input)
        return user_input
    else:
        print("No contract choosen.")

def get_user_input_master():
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    
    # Get user input using a simple dialog
    user_input = simpledialog.askstring("User Input", "Which master are you hiring now?")
    
    if user_input is not None:
        print("Master:", user_input)
        return user_input
    else:
        print("No master choosen.")


def remove_brackets(input_string):
    """
    Remove all bracketed content (square, round, and Chinese-style) from the input string.
    """
    # Remove content in square brackets
    output_string = re.sub(r'\[[^\]]*\]', '', input_string)
    
    # Remove content in round brackets
    output_string = re.sub(r'\([^\)]*\)', '', output_string)
    
    # Remove content in Chinese-style brackets
    output_string = re.sub(r'（[^）]*）', '', output_string)
    
    return output_string

def remove_brackets_in_excel(input_file, output_file):
    """
    Remove all bracketed content from an Excel file.
    """
    # Load the workbook
    workbook = openpyxl.load_workbook(input_file)
    worksheet = workbook.active
    
    # Iterate through each cell and remove bracketed content
    for row in range(1, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row, column=col)
            if cell.value:
                new_value = remove_brackets(str(cell.value))
                new_value = new_value.strip()
                cell.value = new_value
    
    # Save the modified workbook
    workbook.save(output_file)

def remove_bracketed_content(input_file):
   # Prompt the user to confirm the action
   root = tk.Tk()
   root.withdraw()
   confirm = messagebox.askyesno("Confirm Action", "Do you want to remove all the bracketed content from the Excel file?")

   if confirm:
      # Example usage
      remove_brackets_in_excel(f'{input_file}.xlsx', f'{input_file}_b.xlsx')
      return f'{input_file}_b.xlsx'
   else:
      print("Action canceled.")
      return f'{input_file}.xlsx'



contract = remove_bracketed_content(get_user_input())
master = get_user_input_master()
# Read the data from the two Excel files
df_items = pd.read_excel(f'{contract}', header=11)
df_prices = pd.read_excel('EVA_LIBRARY.xlsx', sheet_name=f'{master}')

# Initialize an empty list to store matched prices
matched_prices = []

# Loop through the items in Sheet1
for item in df_items['項目']:
    if item in df_prices['項目'].values:
        # Get the price for the matched item
        price = df_prices.loc[df_prices['項目'] == item, '單價'].values[0]
        matched_prices.append(price)
    else:
        matched_prices.append('')
                

# Add the matched prices as a new column in Sheet1
df_items['單價'] = matched_prices

# Select only the columns '項目' and '單價' for the final output
df_final = df_items[['項目', '位置', '面積/尺寸/數量', '單價']]

contract = contract.replace('.xlsx', '_in.xlsx')

# Write the updated data back to Sheet1 with only '項目' and '單價' columns
df_final.to_excel(f'{contract}', index=False, columns=[ '項目', '位置', '面積/尺寸/數量', '單價'], header=True, startrow=11)

# Load the Excel file
df = pd.read_excel(f'{contract}')

# Insert a new blank column filled with empty values
df.insert(0, '', '')

# Save the modified DataFrame back to an Excel file
df.to_excel(f'{contract}', index=False)