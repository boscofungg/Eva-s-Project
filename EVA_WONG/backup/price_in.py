import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from tkinter import simpledialog
import re
import pandas as pd
import openpyxl
import os

import tkinter as tk
from tkinter import ttk

def perform_calculation(filename):
    filename = filename.replace('.xlsx', '')
    file = filename

    def mm_to_feet(mm):
        feet = mm / 304.8
        return feet

    # Define a custom rounding function
    def round_to_nearest_5(x):
        return round(x, -1)

    df_items = pd.read_excel(f'{file}.xlsx', header = 11)
    i = 0

    # this is for the calculation of 清拆工程 
    while df_items['項目'][i] != "油漆工程":
        i += 1

    i += 1


    # this is for the calculation of 油漆工程 
    workbook = openpyxl.load_workbook(f'{file}.xlsx')

    # Select the specific sheet where you want to add the data
    sheet = workbook['Sheet1']


    while df_items['項目'][i] != "泥水工程":
        if pd.isnull(df_items['面積/尺寸/數量'][i]) or pd.isnull(df_items['單價'][i]) or  df_items['單價'][i] == "1單":
            i+=1  # Checking for NaN values
            continue  # Skip the row if it's empty
        else:
            print(df_items['項目'][i])
            items = df_items['面積/尺寸/數量'][i]
            price = df_items['單價'][i]
            # Input string
            input_string = str(df_items['面積/尺寸/數量'][i])

            # Regular expression to extract numbers
            numbers = re.findall(r'\d+', input_string)

            second_string = str(df_items['單價'][i])
            price = re.findall(r'\d+', second_string)
            numbers = [int(num) for num in numbers]
            price = [int(pri) for pri in price]
            total = numbers[0]*price[0]
            i+=1
            # Write the new data to the specified cell
            total = round_to_nearest_5(total)
            sheet.cell(row = i+12, column=6, value=total)

    i += 1

    # this is for the calculation of 泥水工程 
    while df_items['項目'][i] != "天花":
        if pd.isnull(df_items['面積/尺寸/數量'][i]) or pd.isnull(df_items['單價'][i]) or  df_items['單價'][i] == "1單":
            i+=1  # Checking for NaN values
            continue  # Skip the row if it's empty
        else:
            print(df_items['項目'][i])
            items = df_items['面積/尺寸/數量'][i]
            price = df_items['單價'][i]
            # Input string
            input_string = df_items['面積/尺寸/數量'][i]

            # Regular expression to extract numbers
            numbers = re.findall(r'\d+', input_string)

            second_string = df_items['單價'][i]
            price = re.findall(r'\d+', second_string)
            numbers = [int(num) for num in numbers]
            price = [int(pri) for pri in price]
            total = numbers[0]*price[0]
            i+=1
            # Write the new data to the specified cell
            total = round_to_nearest_5(total)
            sheet.cell(row = i+12, column=6, value=total)


    i+=1

    while df_items['項目'][i] != "門":
        if pd.isnull(df_items['項目'][i]) or pd.isnull(df_items['面積/尺寸/數量'][i]) or pd.isnull(df_items['單價'][i]) or  df_items['單價'][i] == "1單":
            i+=1  # Checking for NaN values
            continue  # Skip the row if it's empty
        else:
            if df_items['項目'][i].find("生口") != -1 and df_items['面積/尺寸/數量'][i].find("生口") != -1: 
                # Input string
                input_string = df_items['面積/尺寸/數量'][i]

                # Regular expression to extract numbers
                numbers = re.findall(r'\d+', input_string)

                # Convert the extracted numbers to integers
                numbers = [int(num) for num in numbers]
                numbers[0] = mm_to_feet(numbers[0])
                numbers[1] = mm_to_feet(numbers[1])
                total_price = numbers[0] * numbers[1]
                # Print the extracted numbers

                second_string = df_items['單價'][i]
                price = re.findall(r'\d+', second_string)
                price = [int(pri) for pri in price]
                price1 = total_price * price[0]
                price2 = numbers[2] * price[1]
                total = price1 + price2
                print(input_string)
                print(second_string)
            elif  df_items['面積/尺寸/數量'][i].find("平方尺") != -1:
                input_string = df_items['面積/尺寸/數量'][i]

                # Regular expression to extract numbers
                numbers = re.findall(r'\d+', input_string)

                # Convert the extracted numbers to integers
                numbers = [int(num) for num in numbers]

                second_string = df_items['單價'][i]

                price = re.findall(r'\d+', second_string)
                price = [int(pri) for pri in price]
                
                total = numbers[0] * price[0]   
            else:
                # Input string
                input_string = df_items['面積/尺寸/數量'][i]

                # Regular expression to extract numbers
                numbers = re.findall(r'\d+', input_string)

                # Convert the extracted numbers to integers
                numbers = [int(num) for num in numbers]
                numbers[0] = mm_to_feet(numbers[0])
                # Print the extracted numbers

                second_string = df_items['單價'][i]
                price = re.findall(r'\d+', second_string)
                price = [int(pri) for pri in price]
                total = numbers[0] * price[0]

            # Write the new data to the specified cell
            i+=1
            total = round_to_nearest_5(total)
            sheet.cell(row = i+12, column=6, value=total)
    

    i+=1

    while df_items['項目'][i] != "木工":
        if pd.isnull(df_items['項目'][i]) or pd.isnull(df_items['面積/尺寸/數量'][i]) or pd.isnull(df_items['單價'][i]) or  df_items['單價'][i] == "1單":
            i+=1  # Checking for NaN values
            continue  # Skip the row if it's empty
        else:
            print(df_items['項目'][i])
            items = df_items['面積/尺寸/數量'][i]
            price = df_items['單價'][i]
            i+=1 
            n=0
            for _ in range(len(items)):
                if items[n].isdigit():
                    n += 1
                else:
                    break
            m = 0
            for _ in range(len(price)):
                if price[m].isdigit():
                    m += 1
                else:
                    break
            price = price[:m]
            price = int(price)
            items = items[:n]
            items = int(items)
            total = items * price

            # Write the new data to the specified cell
            total = round_to_nearest_5(total)
            sheet.cell(row = i+12, column=6, value=total)

    i+=1

    while df_items['項目'][i] != "電燈工程":
        if pd.isnull(df_items['項目'][i]) or pd.isnull(df_items['面積/尺寸/數量'][i]) or pd.isnull(df_items['單價'][i]) or  df_items['單價'][i] == "1單":
            i+=1  # Checking for NaN values
            continue  # Skip the row if it's empty
        else:
            if df_items['單價'][i].find("圓角") != -1:
                input_string = df_items['面積/尺寸/數量'][i]
                second_string = df_items['單價'][i]
                i+=1
                numbers = re.findall(r'\d+', input_string)
                numbers = [int(num) for num in numbers]
                numbers[0] = mm_to_feet(numbers[0])

                price = re.findall(r'\d+', second_string)
                price = [int(pri) for pri in price]
                total = numbers[0] * price[0] + price[1]
            elif df_items['面積/尺寸/數量'][i].find("長") != -1 and df_items['面積/尺寸/數量'][i].find("高") != -1 and df_items['單價'][i].find("圓角") == -1:
                input_string = df_items['面積/尺寸/數量'][i]
                second_string = df_items['單價'][i]
                i+=1
                numbers = re.findall(r'\d+', input_string)
                numbers = [int(num) for num in numbers]
                numbers[0] = mm_to_feet(numbers[0])
                price = re.findall(r'\d+', second_string)
                price = [int(pri) for pri in price]
                total = numbers[0] * price[0]
            else:
                input_string = df_items['面積/尺寸/數量'][i]
                second_string = df_items['單價'][i]
                i+=1
                numbers = re.findall(r'\d+', input_string)
                numbers = [int(num) for num in numbers]
                price = re.findall(r'\d+', second_string)
                price = [int(pri) for pri in price]
                total = numbers[0] * price[0]
            # Write the new data to the specified cell
            total = round_to_nearest_5(total)
            sheet.cell(row = i+12, column=6, value=total)

    i+=1

    while df_items['項目'][i] != "附註":
        if pd.isnull(df_items['項目'][i]) or pd.isnull(df_items['面積/尺寸/數量'][i]) or pd.isnull(df_items['單價'][i]) or  df_items['單價'][i] == "1單":
            i+=1  # Checking for NaN values
            continue  # Skip the row if it's empty
        elif df_items['面積/尺寸/數量'][i] == '一單' or df_items['單價'][i] == '一單':
            i+=1
            continue
        else:
            print(df_items['項目'][i])
            items = df_items['面積/尺寸/數量'][i]
            price = df_items['單價'][i]
            i+=1 
            numbers = re.findall(r'\d+', items)
            numbers = [int(num) for num in numbers]
            price = re.findall(r'\d+', price)
            price = [int(pri) for pri in price]
            total = numbers[0] * price[0]
            total = round_to_nearest_5(total)
            # Write the new data to the specified cell
            sheet.cell(row = i+12, column=6, value=total)

    # Save the updated Excel file
    workbook.save(f'{file}.xlsx')
    print("Done!")

def main_window(input_file):
   # Prompt the user to confirm the action
   root = tk.Tk()
   root.withdraw()
   confirm = messagebox.askyesno("Confirm Action", "Do you want to do the calculations for you?")

   if confirm:
      # Example usage
        perform_calculation(input_file)
   else:
        print("Action canceled.")


def get_user_input():
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    
    # Get user input using a simple dialog
    user_input = simpledialog.askstring("User Input", "Which contract are you working on now?")
    
    if user_input is not None:
        print("Contract:", user_input)
        return f'{user_input}'
    else:
        print("No contract choosen.")

def get_user_input_master():
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    
    # Get user input using a simple dialog
    user_input = simpledialog.askstring("User Input", "Which master are you hiring now?")
    
    if user_input is not None:
        print("Master:", user_input)
        return user_input
    else:
        print("No master choosen.")


def remove_brackets(input_string):
    """
    Remove all bracketed content (square, round, and Chinese-style) from the input string.
    """
    # Remove content in square brackets
    output_string = re.sub(r'\[[^\]]*\]', '', input_string)
    
    # Remove content in round brackets
    output_string = re.sub(r'\([^\)]*\)', '', output_string)
    
    # Remove content in Chinese-style brackets
    output_string = re.sub(r'（[^）]*）', '', output_string)
    
    return output_string

def remove_brackets_in_excel(input_file, output_file):
    """
    Remove all bracketed content from an Excel file.
    """
    # Load the workbook
    workbook = openpyxl.load_workbook(input_file)
    worksheet = workbook.active
    
    # Iterate through each cell and remove bracketed content
    for row in range(1, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row, column=col)
            if cell.value:
                new_value = remove_brackets(str(cell.value))
                new_value = new_value.strip()
                cell.value = new_value
    
    # Save the modified workbook
    workbook.save(output_file)

def remove_bracketed_content(input_file):
   # Prompt the user to confirm the action
   root = tk.Tk()
   root.withdraw()
   confirm = messagebox.askyesno("Confirm Action", "Do you want to remove all the bracketed content from the Excel file?")

   if confirm:
      # Example usage
      remove_brackets_in_excel(f'{input_file}.xlsx', f'{input_file}_b.xlsx')
      return f'{input_file}_b.xlsx'
   else:
      print("Action canceled.")
      return f'{input_file}.xlsx'

def delete_excel_file(filename):
    # Get the current directory
    current_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(current_dir, filename)
    os.remove(file_path)

if __name__ == "__main__":
    contract = remove_bracketed_content(get_user_input())
    master = get_user_input_master()
    # Read the data from the two Excel files
    df_items = pd.read_excel(f'{contract}', header=11)
    df_prices = pd.read_excel('EVA_LIBRARY.xlsx', sheet_name=f'{master}')

    # Initialize an empty list to store matched prices
    matched_prices = []

    # Loop through the items in Sheet1
    for item in df_items['項目']:
        if item in df_prices['項目'].values:
            # Get the price for the matched item
            price = df_prices.loc[df_prices['項目'] == item, '單價'].values[0]
            matched_prices.append(price)
        else:
            matched_prices.append('')
                    

    # Add the matched prices as a new column in Sheet1
    df_items['單價'] = matched_prices

    # Select only the columns '項目' and '單價' for the final output
    df_final = df_items[['項目', '位置', '面積/尺寸/數量', '單價']]

    new_contract = contract.replace('.xlsx', '_in.xlsx')

    # Write the updated data back to Sheet1 with only '項目' and '單價' columns
    df_final.to_excel(f'{new_contract}', index=False, columns=[ '項目', '位置', '面積/尺寸/數量', '單價'], header=True, startrow=11)

    # Load the Excel file
    df = pd.read_excel(f'{new_contract}')

    # Insert a new blank column filled with empty values
    df.insert(0, '', '')

    # Save the modified DataFrame back to an Excel file
    df.to_excel(f'{new_contract}', index=False)

    if contract.find('_b') != -1:
        delete_excel_file(contract)

    main_window(new_contract)