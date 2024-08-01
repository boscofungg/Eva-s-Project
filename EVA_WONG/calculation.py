import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from tkinter import simpledialog
import re
import pandas as pd
import openpyxl
import re


def get_user_input():
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    
    # Get user input using a simple dialog
    user_input = simpledialog.askstring("User Input", "Which excel do you want to calculate?")
    
    if user_input is not None:
        print("Excel:", user_input)
        return user_input
    else:
        print("No contract choosen.")

def mm_to_feet(mm):
    feet = mm / 304.8
    return feet

# Define a custom rounding function
def round_to_nearest_5(x):
    return round(x, -1)


file = get_user_input()

df_items = pd.read_excel(f'{file}.xlsx', header = 11)
i = 0

# this is for the calculation of 清拆工程 
while df_items['項目'][i] != "油漆工程":
    i += 1

i += 1


# this is for the calculation of 油漆工程 
workbook = openpyxl.load_workbook(f'{file}.xlsx')

# Select the specific sheet where you want to add the data
sheet = workbook['Sheet1']


while df_items['項目'][i] != "泥水工程":
    if pd.isnull(df_items['面積/尺寸/數量'][i]) or pd.isnull(df_items['單價'][i]) or  df_items['單價'][i] == "1單":
        i+=1  # Checking for NaN values
        continue  # Skip the row if it's empty
    else:
        print(df_items['項目'][i])
        items = df_items['面積/尺寸/數量'][i]
        price = df_items['單價'][i]
        # Input string
        input_string = str(df_items['面積/尺寸/數量'][i])

        # Regular expression to extract numbers
        numbers = re.findall(r'\d+', input_string)

        second_string = str(df_items['單價'][i])
        price = re.findall(r'\d+', second_string)
        numbers = [int(num) for num in numbers]
        price = [int(pri) for pri in price]
        total = numbers[0]*price[0]
        i+=1
        # Write the new data to the specified cell
        total = round_to_nearest_5(total)
        sheet.cell(row = i+12, column=6, value=total)

i += 1

# this is for the calculation of 泥水工程 
while df_items['項目'][i] != "天花":
    if pd.isnull(df_items['面積/尺寸/數量'][i]) or pd.isnull(df_items['單價'][i]) or  df_items['單價'][i] == "1單":
        i+=1  # Checking for NaN values
        continue  # Skip the row if it's empty
    else:
        print(df_items['項目'][i])
        items = df_items['面積/尺寸/數量'][i]
        price = df_items['單價'][i]
        # Input string
        input_string = df_items['面積/尺寸/數量'][i]

        # Regular expression to extract numbers
        numbers = re.findall(r'\d+', input_string)

        second_string = df_items['單價'][i]
        price = re.findall(r'\d+', second_string)
        numbers = [int(num) for num in numbers]
        price = [int(pri) for pri in price]
        total = numbers[0]*price[0]
        i+=1
        # Write the new data to the specified cell
        total = round_to_nearest_5(total)
        sheet.cell(row = i+12, column=6, value=total)


i+=1

while df_items['項目'][i] != "門":
    if pd.isnull(df_items['項目'][i]) or pd.isnull(df_items['面積/尺寸/數量'][i]) or pd.isnull(df_items['單價'][i]) or  df_items['單價'][i] == "1單":
        i+=1  # Checking for NaN values
        continue  # Skip the row if it's empty
    else:
        if df_items['項目'][i].find("生口") != -1 and df_items['面積/尺寸/數量'][i].find("生口") != -1: 
            # Input string
            input_string = df_items['面積/尺寸/數量'][i]

            # Regular expression to extract numbers
            numbers = re.findall(r'\d+', input_string)

            # Convert the extracted numbers to integers
            numbers = [int(num) for num in numbers]
            numbers[0] = mm_to_feet(numbers[0])
            numbers[1] = mm_to_feet(numbers[1])
            total_price = numbers[0] * numbers[1]
            # Print the extracted numbers

            second_string = df_items['單價'][i]
            price = re.findall(r'\d+', second_string)
            price = [int(pri) for pri in price]
            price1 = total_price * price[0]
            price2 = numbers[2] * price[1]
            total = price1 + price2
            print(input_string)
            print(second_string)
        elif  df_items['面積/尺寸/數量'][i].find("平方尺") != -1:
            input_string = df_items['面積/尺寸/數量'][i]

            # Regular expression to extract numbers
            numbers = re.findall(r'\d+', input_string)

            # Convert the extracted numbers to integers
            numbers = [int(num) for num in numbers]

            second_string = df_items['單價'][i]

            price = re.findall(r'\d+', second_string)
            price = [int(pri) for pri in price]
            
            total = numbers[0] * price[0]   
        else:
            # Input string
            input_string = df_items['面積/尺寸/數量'][i]

            # Regular expression to extract numbers
            numbers = re.findall(r'\d+', input_string)

            # Convert the extracted numbers to integers
            numbers = [int(num) for num in numbers]
            numbers[0] = mm_to_feet(numbers[0])
            # Print the extracted numbers

            second_string = df_items['單價'][i]
            price = re.findall(r'\d+', second_string)
            price = [int(pri) for pri in price]
            total = numbers[0] * price[0]

        # Write the new data to the specified cell
        i+=1
        total = round_to_nearest_5(total)
        sheet.cell(row = i+12, column=6, value=total)
  

i+=1

while df_items['項目'][i] != "木工":
    if pd.isnull(df_items['項目'][i]) or pd.isnull(df_items['面積/尺寸/數量'][i]) or pd.isnull(df_items['單價'][i]) or  df_items['單價'][i] == "1單":
        i+=1  # Checking for NaN values
        continue  # Skip the row if it's empty
    else:
        print(df_items['項目'][i])
        items = df_items['面積/尺寸/數量'][i]
        price = df_items['單價'][i]
        i+=1 
        n=0
        for _ in range(len(items)):
            if items[n].isdigit():
                n += 1
            else:
                break
        m = 0
        for _ in range(len(price)):
            if price[m].isdigit():
                m += 1
            else:
                break
        price = price[:m]
        price = int(price)
        items = items[:n]
        items = int(items)
        total = items * price

        # Write the new data to the specified cell
        total = round_to_nearest_5(total)
        sheet.cell(row = i+12, column=6, value=total)

i+=1

while df_items['項目'][i] != "電燈工程":
    if pd.isnull(df_items['項目'][i]) or pd.isnull(df_items['面積/尺寸/數量'][i]) or pd.isnull(df_items['單價'][i]) or  df_items['單價'][i] == "1單":
        i+=1  # Checking for NaN values
        continue  # Skip the row if it's empty
    else:
        if df_items['單價'][i].find("圓角") != -1:
            input_string = df_items['面積/尺寸/數量'][i]
            second_string = df_items['單價'][i]
            i+=1
            numbers = re.findall(r'\d+', input_string)
            numbers = [int(num) for num in numbers]
            numbers[0] = mm_to_feet(numbers[0])

            price = re.findall(r'\d+', second_string)
            price = [int(pri) for pri in price]
            total = numbers[0] * price[0] + price[1]
        elif df_items['面積/尺寸/數量'][i].find("長") != -1 or df_items['面積/尺寸/數量'][i].find("高") != -1: 
            input_string = df_items['面積/尺寸/數量'][i]
            second_string = df_items['單價'][i]
            i+=1
            numbers = re.findall(r'\d+', input_string)
            numbers = [int(num) for num in numbers]
            numbers[0] = mm_to_feet(numbers[0])
            price = re.findall(r'\d+', second_string)
            price = [int(pri) for pri in price]
            total = numbers[0] * price[0]
        else:
            input_string = df_items['面積/尺寸/數量'][i]
            second_string = df_items['單價'][i]
            i+=1
            numbers = re.findall(r'\d+', input_string)
            numbers = [int(num) for num in numbers]
            price = re.findall(r'\d+', second_string)
            price = [int(pri) for pri in price]
            total = numbers[0] * price[0]
        # Write the new data to the specified cell
        total = round_to_nearest_5(total)
        sheet.cell(row = i+12, column=6, value=total)

i+=1

while df_items['項目'][i] != "附註":
    if pd.isnull(df_items['項目'][i]) or pd.isnull(df_items['面積/尺寸/數量'][i]) or pd.isnull(df_items['單價'][i]) or  df_items['單價'][i] == "1單":
        i+=1  # Checking for NaN values
        continue  # Skip the row if it's empty
    elif df_items['面積/尺寸/數量'][i] == '一單' or df_items['單價'][i] == '一單':
        i+=1
        continue
    else:
        print(df_items['項目'][i])
        items = df_items['面積/尺寸/數量'][i]
        price = df_items['單價'][i]
        i+=1 
        numbers = re.findall(r'\d+', items)
        numbers = [int(num) for num in numbers]
        price = re.findall(r'\d+', price)
        price = [int(pri) for pri in price]
        total = numbers[0] * price[0]
        total = round_to_nearest_5(total)
        # Write the new data to the specified cell
        sheet.cell(row = i+12, column=6, value=total)

# Save the updated Excel file
workbook.save(f'{file}.xlsx')
print("Done!")